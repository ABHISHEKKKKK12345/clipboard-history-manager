"""
╔══════════════════════════════════════════════════════════════════════════════╗
║          Clipboard History Manager  ·  v1.0.0  ·  Production-Ready           ║
║          Cross-Platform  |  Thread-Safe  |  Lifecycle-Safe                   ║
╠══════════════════════════════════════════════════════════════════════════════╣
║  Author   : Abhishek Srivastava                                              ║
║  Version  : 1.0.0                                                            ║
║  License  : MIT                                                              ║
║  Python   : 3.9+                                                             ║
║  Requires : customtkinter>=5.2, pyperclip, openpyxl                          ║
║  Optional : pynput  (global hotkey  Ctrl+Shift+V)                            ║
╚══════════════════════════════════════════════════════════════════════════════╝
"""

from __future__ import annotations

# ─────────────────────────────────────────────────────────────────────────────
# Standard Library
# ─────────────────────────────────────────────────────────────────────────────
import json
import logging
import os
import platform
import queue
import re
import sys
import threading
import time
import traceback
from dataclasses import asdict, dataclass, field
from datetime import datetime
from logging.handlers import RotatingFileHandler
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Set, Tuple

# ─────────────────────────────────────────────────────────────────────────────
# Third-party — hard requirements
# ─────────────────────────────────────────────────────────────────────────────
try:
    import customtkinter as ctk
except ImportError:
    sys.exit(
        "\n  ERROR: customtkinter is not installed.\n"
        "  Fix  :  pip install customtkinter\n"
    )

try:
    import pyperclip
except ImportError:
    sys.exit(
        "\n  ERROR: pyperclip is not installed.\n"
        "  Fix  :  pip install pyperclip\n"
    )

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.exceptions import IllegalCharacterError as _OpenpyxlIllegalChar
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False
    _OpenpyxlIllegalChar = None  # type: ignore[assignment,misc]

# ─────────────────────────────────────────────────────────────────────────────
# Third-party — optional
# ─────────────────────────────────────────────────────────────────────────────
try:
    from pynput import keyboard as _pynput_keyboard
    PYNPUT_OK = True
except ImportError:
    _pynput_keyboard = None  # type: ignore[assignment]
    PYNPUT_OK = False

# ─────────────────────────────────────────────────────────────────────────────
# Application constants
# ─────────────────────────────────────────────────────────────────────────────
APP_NAME         : str  = "Clipboard History Manager"
APP_VERSION      : str  = "1.0.0"
APP_AUTHOR       : str  = "Abhishek Srivastava"

APP_DATA_DIR     : Path = Path.home() / ".clipboard_manager"
DB_FILE          : Path = APP_DATA_DIR / "history.json"
LOG_FILE         : Path = APP_DATA_DIR / "app.log"
SETTINGS_FILE    : Path = APP_DATA_DIR / "settings.json"

# Storage limits
MAX_ITEMS_MIN    : int  = 10
MAX_ITEMS_MAX    : int  = 2_000
MAX_ITEM_BYTES   : int  = 1_000_000      # 1 MB hard limit per clipboard entry

# Timing
POLL_INTERVAL_MS : int  = 400            # clipboard polling cadence (ms)
DEBOUNCE_SEC     : float = 0.30          # minimum gap between identical clips
SEARCH_DEBOUNCE  : int  = 120            # ms delay before search filter fires
WATCHDOG_MS      : int  = 5_000          # monitor health-check interval (ms)
PASTE_TIMEOUT_S  : float = 1.0           # max seconds for pyperclip.paste()
COPY_TIMEOUT_S   : float = 1.0           # max seconds for pyperclip.copy()
KIND_TEXT_LIMIT  : int  = 2048           # max chars checked for kind detection (FIX-H)

# UI virtualisation
VIRTUAL_PAGE     : int  = 80            # maximum ClipCards rendered at once
QUEUE_DRAIN_MAX  : int  = 100           # maximum queue messages per UI tick

# Shutdown polling
_SHUTDOWN_POLL_MS : int = 50            # how often main thread checks store-done event

# ── XLSX export row / column geometry ──────────────────────────────────────
XL_ROW_HEADER            : float = 18.0
XL_ROW_DATA              : float = 16.0
XL_ROW_CONTENT_PER_LINE  : float = 15.0
XL_ROW_CONTENT_MAX_LINES : int   = 6

# ─────────────────────────────────────────────────────────────────────────────
# Colour palette
# ─────────────────────────────────────────────────────────────────────────────
THEME: Dict[str, str] = {
    "bg"         : "#0D0F14",
    "panel"      : "#13161E",
    "card"       : "#1A1D27",
    "card_hover" : "#22263A",
    "card_pin"   : "#1B2340",
    "border"     : "#2A2D3E",
    "accent"     : "#5B7FFF",
    "accent2"    : "#A78BFA",
    "success"    : "#34D399",
    "warn"       : "#FBBF24",
    "danger"     : "#F87171",
    "text"       : "#E2E8F0",
    "text_dim"   : "#64748B",
    "text_muted" : "#3D4460",
    "scrollbar"  : "#2A2D3E",
}

# ─────────────────────────────────────────────────────────────────────────────
# Logging  (singleton guard prevents duplicate handlers on re-import)
# ─────────────────────────────────────────────────────────────────────────────

def _setup_logging() -> logging.Logger:
    logger = logging.getLogger("ClipMgr")
    if logger.handlers:
        return logger
    logger.setLevel(logging.DEBUG)

    fmt = logging.Formatter(
        "%(asctime)s [%(levelname)-8s] %(name)s — %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )
    try:
        APP_DATA_DIR.mkdir(parents=True, exist_ok=True)
        fh = RotatingFileHandler(
            LOG_FILE, maxBytes=5_000_000, backupCount=3, encoding="utf-8"
        )
        fh.setFormatter(fmt)
        fh.setLevel(logging.DEBUG)
        logger.addHandler(fh)
    except OSError as exc:
        print(f"[ClipMgr] WARNING: could not open log file: {exc}", file=sys.stderr)

    sh = logging.StreamHandler(sys.stdout)
    sh.setFormatter(fmt)
    sh.setLevel(logging.INFO)
    logger.addHandler(sh)
    return logger


log: logging.Logger = _setup_logging()


# ─────────────────────────────────────────────────────────────────────────────
# XML sanitisation helper
# ─────────────────────────────────────────────────────────────────────────────
_SURROGATE_RE: re.Pattern = re.compile(r"[\ud800-\udfff]")
_ILLEGAL_XML_RE: re.Pattern = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")


def _sanitize_for_xml(value: object) -> str:
    """
    Convert *value* to str, strip lone Unicode surrogates and XML-1.0-illegal
    control characters. MUST be called on every string before writing to an
    openpyxl cell.
    """
    if value is None:
        return ""
    text = str(value)
    text = _SURROGATE_RE.sub("", text)
    text = text.encode("utf-8", errors="replace").decode("utf-8", errors="replace")
    return _ILLEGAL_XML_RE.sub("", text)


# ─────────────────────────────────────────────────────────────────────────────
# Timeout-safe clipboard paste
# ─────────────────────────────────────────────────────────────────────────────

_paste_semaphore: threading.BoundedSemaphore = threading.BoundedSemaphore(1)


def _paste_with_timeout(timeout: float = PASTE_TIMEOUT_S) -> str:
    """
    Call pyperclip.paste() from a daemon thread with a hard timeout.
    Uses _paste_semaphore to guarantee at most one paste thread at a time.
    Raises TimeoutError if semaphore is unavailable or if paste() blocks.
    Re-raises any exception thrown by pyperclip.paste() itself.
    """
    if not _paste_semaphore.acquire(blocking=False):
        raise TimeoutError("pyperclip.paste() skipped — previous call still in flight")

    result: List[str] = [""]
    exc_holder: List[Optional[BaseException]] = [None]

    def _worker() -> None:
        try:
            result[0] = pyperclip.paste() or ""
        except Exception as exc:  # noqa: BLE001
            exc_holder[0] = exc
        finally:
            _paste_semaphore.release()

    t = threading.Thread(target=_worker, daemon=True, name="ClipMgr/PasteWorker")
    t.start()
    t.join(timeout=timeout)
    if t.is_alive():
        # _worker will release the semaphore when it eventually unblocks
        raise TimeoutError(
            f"pyperclip.paste() did not return within {timeout:.1f}s"
        )
    if exc_holder[0] is not None:
        raise exc_holder[0]  # type: ignore[misc]
    return result[0]


# ─────────────────────────────────────────────────────────────────────────────
# Timeout-safe clipboard copy  (FIX-B: added _copy_semaphore)
# ─────────────────────────────────────────────────────────────────────────────

# FIX-B: Bounded semaphore prevents concurrent copy threads from accumulating.
_copy_semaphore: threading.BoundedSemaphore = threading.BoundedSemaphore(1)


def _copy_with_timeout(text: str, timeout: float = COPY_TIMEOUT_S) -> None:
    """
    Call pyperclip.copy(text) from a daemon thread with a hard timeout.
    Uses _copy_semaphore to guarantee at most one copy thread at a time.
    Raises TimeoutError if semaphore is unavailable or if copy() blocks.
    Re-raises any exception thrown by pyperclip.copy() itself.
    """
    # FIX-B: If a copy is already in progress, skip this call.
    if not _copy_semaphore.acquire(blocking=False):
        raise TimeoutError("pyperclip.copy() skipped — previous call still in flight")

    exc_holder: List[Optional[BaseException]] = [None]

    def _worker() -> None:
        try:
            pyperclip.copy(text)
        except Exception as exc:  # noqa: BLE001
            exc_holder[0] = exc
        finally:
            _copy_semaphore.release()

    t = threading.Thread(target=_worker, daemon=True, name="ClipMgr/CopyWorker")
    t.start()
    t.join(timeout=timeout)
    if t.is_alive():
        # _worker will release the semaphore when it eventually unblocks
        raise TimeoutError(
            f"pyperclip.copy() did not complete within {timeout:.1f}s"
        )
    if exc_holder[0] is not None:
        raise exc_holder[0]  # type: ignore[misc]


# ═════════════════════════════════════════════════════════════════════════════
# DATA MODEL
# ═════════════════════════════════════════════════════════════════════════════

@dataclass
class ClipItem:
    """Represents a single clipboard history entry."""

    id        : str
    text      : str
    timestamp : float
    pinned    : bool      = False
    tags      : List[str] = field(default_factory=list)

    @property
    def preview(self) -> str:
        """Single-line truncated preview (≤ 120 chars)."""
        s = " ".join(self.text.split())
        return (s[:119] + "…") if len(s) > 120 else s

    @property
    def dt(self) -> datetime:
        return datetime.fromtimestamp(self.timestamp)

    @property
    def human_time(self) -> str:
        delta = time.time() - self.timestamp
        if delta < 60:
            return f"{int(delta)}s ago"
        if delta < 3_600:
            return f"{int(delta / 60)}m ago"
        if delta < 86_400:
            return f"{int(delta / 3_600)}h ago"
        return self.dt.strftime("%d %b %Y")

    @property
    def kind(self) -> str:
        """
        Classify the clipboard entry.
        FIX-H: Bails out early for large text; regex runs on a capped slice.
        """
        t = self.text.strip()
        # FIX-H: skip expensive regex on very large text
        if len(t) > KIND_TEXT_LIMIT:
            return "text"
        # Use a safe slice for regex matching
        sample = t[:KIND_TEXT_LIMIT]
        if re.match(
            r"^https?://[^\s/$.?#][^\s]*\.[^\s]{2,}$", sample, re.IGNORECASE
        ):
            return "url"
        if re.match(r"^[\w.%+\-]+@[\w\-]+(?:\.[\w\-]+)+$", sample, re.IGNORECASE):
            return "email"
        digits = re.sub(r"[\s\-+().]+", "", sample)
        if digits.isdigit() and 7 <= len(digits) <= 15:
            return "phone"
        if "\n" in t and len(t) > 80 and re.search(r"[{}\[\]();:=]", sample):
            return "code"
        return "text"

    def to_dict(self) -> dict:
        return asdict(self)

    @classmethod
    def from_dict(cls, d: dict) -> "ClipItem":
        return cls(
            id        = str(d.get("id", "")),
            text      = str(d.get("text", "")),
            timestamp = float(d.get("timestamp", time.time())),
            pinned    = bool(d.get("pinned", False)),
            tags      = list(d.get("tags", [])),
        )

    @classmethod
    def new(cls, text: str) -> "ClipItem":
        return cls(id=str(time.time_ns()), text=text, timestamp=time.time())


# ═════════════════════════════════════════════════════════════════════════════
# SETTINGS
# ═════════════════════════════════════════════════════════════════════════════

_SETTINGS_SCHEMA: Dict[str, type] = {
    "max_items"       : int,
    "dedup"           : bool,
    "trim_whitespace" : bool,
    "start_minimized" : bool,
    "font_size"       : int,
}

FONT_SIZE_MIN : int = 10
FONT_SIZE_MAX : int = 20


@dataclass
class Settings:
    """Persistent user preferences, loaded from / saved to JSON."""

    max_items       : int  = 500
    dedup           : bool = True
    trim_whitespace : bool = True
    start_minimized : bool = False
    font_size       : int  = 13

    def save(self) -> None:
        try:
            APP_DATA_DIR.mkdir(parents=True, exist_ok=True)
            tmp = SETTINGS_FILE.with_suffix(".tmp")
            tmp.write_text(json.dumps(asdict(self), indent=2), encoding="utf-8")
            os.replace(tmp, SETTINGS_FILE)
        except OSError as exc:
            log.error("Settings save failed: %s", exc)

    @classmethod
    def load(cls) -> "Settings":
        try:
            raw = json.loads(SETTINGS_FILE.read_text(encoding="utf-8"))
            s   = cls()

            unknown = set(raw) - set(_SETTINGS_SCHEMA) - {"version"}
            if unknown:
                log.warning(
                    "Settings: unknown keys will be ignored: %s", sorted(unknown)
                )

            for key, cast in _SETTINGS_SCHEMA.items():
                if key not in raw:
                    continue
                val = raw[key]
                try:
                    if cast is bool:
                        if isinstance(val, bool):
                            setattr(s, key, val)
                        elif isinstance(val, str):
                            setattr(s, key, val.lower() not in ("false", "0", "no", ""))
                        else:
                            setattr(s, key, bool(val))
                    else:
                        setattr(s, key, cast(val))
                except (ValueError, TypeError) as exc:
                    log.warning(
                        "Settings key %r ignored (%s) — keeping default", key, exc
                    )
            return s
        except FileNotFoundError:
            return cls()
        except Exception as exc:
            log.warning("Settings file unreadable (%s) — using defaults", exc)
            return cls()


# ═════════════════════════════════════════════════════════════════════════════
# HISTORY STORE  (thread-safe)
# ═════════════════════════════════════════════════════════════════════════════

class HistoryStore:
    """
    Thread-safe, persistent store for ClipItem objects.

    Persistence: generation counter (_dirty_gen) incremented on every mutation.
    Flush thread persists when generation advances. (FIX-J: gen captured under
    lock before event clear to eliminate race.)

    Dedup: _text_set provides O(1) membership tests.
    Load dedup: _load() deduplicates by text; newest timestamp wins.
    delete() safety: only removes text from _text_set if no other item holds it.
    """

    def __init__(self, settings: Settings) -> None:
        self._settings      = settings
        self._lock          = threading.RLock()
        self._items         : List[ClipItem] = []
        self._text_set      : Set[str] = set()

        self._dirty_gen           : int = 0
        self._last_persisted_gen  : int = 0
        self._dirty               = threading.Event()

        self._stop     = threading.Event()
        self._flush_th = threading.Thread(
            target=self._flush_loop, daemon=True, name="ClipMgr/StoreFlush"
        )
        self._load()
        self._flush_th.start()

    # ── public API ────────────────────────────────────────────────────────

    def add(self, text: str) -> Optional[ClipItem]:
        if not text:
            return None
        if self._settings.trim_whitespace:
            text = text.strip()
        if not text:
            return None
        encoded_len = len(text.encode("utf-8", errors="replace"))
        if encoded_len > MAX_ITEM_BYTES:
            log.warning(
                "Clipboard entry rejected — too large (%s bytes, limit %s)",
                f"{encoded_len:,}", f"{MAX_ITEM_BYTES:,}",
            )
            return None

        with self._lock:
            if self._settings.dedup and text in self._text_set:
                for idx, existing in enumerate(self._items):
                    if existing.text == text:
                        self._items.pop(idx)
                        existing.timestamp = time.time()
                        insert_at = (
                            0 if existing.pinned
                            else next(
                                (i for i, x in enumerate(self._items) if not x.pinned),
                                len(self._items),
                            )
                        )
                        self._items.insert(insert_at, existing)
                        self._mark_dirty()
                        return existing

            item      = ClipItem.new(text)
            insert_at = next(
                (i for i, x in enumerate(self._items) if not x.pinned),
                len(self._items),
            )
            self._items.insert(insert_at, item)
            self._text_set.add(text)
            self._prune()
            self._mark_dirty()
            return item

    def all(self) -> List[ClipItem]:
        with self._lock:
            return list(self._items)

    def search(self, query: str) -> List[ClipItem]:
        q = query.casefold()
        with self._lock:
            return [i for i in self._items if q in i.text.casefold()]

    def pin(self, item_id: str) -> None:
        with self._lock:
            for item in self._items:
                if item.id == item_id:
                    item.pinned = not item.pinned
                    self._items.sort(key=lambda x: (0 if x.pinned else 1, -x.timestamp))
                    self._mark_dirty()
                    return

    def delete(self, item_id: str) -> None:
        with self._lock:
            before        = len(self._items)
            deleted_text  : Optional[str] = None

            for item in self._items:
                if item.id == item_id:
                    deleted_text = item.text
                    break

            self._items = [i for i in self._items if i.id != item_id]

            if len(self._items) < before and deleted_text is not None:
                if not any(i.text == deleted_text for i in self._items):
                    self._text_set.discard(deleted_text)
                self._mark_dirty()

    def clear_unpinned(self) -> int:
        with self._lock:
            before      = len(self._items)
            self._items = [i for i in self._items if i.pinned]
            self._text_set = {i.text for i in self._items}
            removed     = before - len(self._items)
            if removed:
                self._mark_dirty()
            return removed

    def clear_all(self) -> None:
        with self._lock:
            self._items.clear()
            self._text_set.clear()
            self._mark_dirty()

    def get(self, item_id: str) -> Optional[ClipItem]:
        with self._lock:
            for item in self._items:
                if item.id == item_id:
                    return item
        return None

    def update_settings(self, settings: Settings) -> None:
        with self._lock:
            self._settings = settings
            self._prune()
            self._mark_dirty()

    # ── Export: XLSX ──────────────────────────────────────────────────────

    def export_xlsx(self, path: Path) -> int:
        """
        Export clipboard history as a formatted Excel workbook.
        Atomic save. FIX-G: explicit IllegalCharacterError catch added.
        """
        if not OPENPYXL_OK:
            raise RuntimeError(
                "openpyxl is required for Excel export.\n"
                "Install it with:  pip install openpyxl"
            )

        items = self.all()

        XL = {
            "hdr_bg"   : "1E3A5F",
            "hdr_fg"   : "FFFFFF",
            "alt_row"  : "EEF4FB",
            "white"    : "FFFFFF",
            "pinned"   : "FFF3CD",
            "total"    : "DBEAFE",
            "meta_bg"  : "F9FAFB",
            "stat_hdr" : "1E3A5F",
        }
        KIND_HEX: Dict[str, str] = {
            "url"  : "2563EB",
            "email": "7C3AED",
            "phone": "059669",
            "code" : "D97706",
            "text" : "374151",
        }
        KIND_LABEL: Dict[str, str] = {
            "url"  : "URL",
            "email": "Email",
            "phone": "Phone",
            "code" : "Code",
            "text" : "Text",
        }
        PINNED_KIND_HEX = "92400E"

        def _fill(hex6: str) -> "PatternFill":
            return PatternFill("solid", start_color=hex6, end_color=hex6)

        def _border() -> "Border":
            edge = Side(style="thin", color="D1D5DB")
            return Border(left=edge, right=edge, top=edge, bottom=edge)

        def _align(h: str = "center", v: str = "center", wrap: bool = False) -> "Alignment":
            return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

        COLS: List[Tuple[str, int, int, str]] = [
            ("#",            8,  10,  "center"),
            ("Item ID",     24,  32,  "center"),
            ("Date & Time", 28,  34,  "center"),
            ("Time Ago",    16,  22,  "center"),
            ("Kind",        14,  18,  "center"),
            ("Pinned",      12,  16,  "center"),
            ("Char Count",  18,  24,  "center"),
            ("Tags",        22,  40,  "center"),
            ("Content",     80, 110,   "left"),
        ]
        N_COLS = len(COLS)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "History"

        HDR_FONT  = Font(name="Arial", bold=True,  size=11, color=XL["hdr_fg"])
        NORM_FONT = Font(name="Arial", bold=False, size=10, color="111827")
        PIN_FONT  = Font(name="Arial", bold=True,  size=10, color=PINNED_KIND_HEX)
        MONO_FONT = Font(name="Courier New", size=9, color="1F2937")

        hdr_fill  = _fill(XL["hdr_bg"])
        hdr_align = _align("center", "center", wrap=False)
        for ci, (label, *_) in enumerate(COLS, 1):
            cell           = ws.cell(row=1, column=ci, value=_sanitize_for_xml(label))
            cell.font      = HDR_FONT
            cell.fill      = hdr_fill
            cell.alignment = hdr_align
            cell.border    = _border()

        ws.row_dimensions[1].height = XL_ROW_HEADER

        for ri, item in enumerate(items, 2):
            is_pin = item.pinned
            row_bg = (
                XL["pinned"]  if is_pin
                else (XL["alt_row"] if ri % 2 == 0 else XL["white"])
            )
            row_fill  = _fill(row_bg)
            base_font = PIN_FONT if is_pin else NORM_FONT

            pin_label = "Yes" if is_pin else "No"

            row_vals: List = [
                ri - 1,
                _sanitize_for_xml(item.id),
                _sanitize_for_xml(item.dt.strftime("%Y-%m-%d  %H:%M:%S")),
                _sanitize_for_xml(item.human_time),
                _sanitize_for_xml(KIND_LABEL.get(item.kind, item.kind.capitalize())),
                _sanitize_for_xml(pin_label),
                len(item.text),
                _sanitize_for_xml("; ".join(item.tags) if item.tags else "—"),
                _sanitize_for_xml(item.text),
            ]

            for ci, value in enumerate(row_vals, 1):
                cell        = ws.cell(row=ri, column=ci, value=value)
                cell.fill   = row_fill
                cell.border = _border()

                if ci == 9:
                    cell.font      = MONO_FONT
                    cell.alignment = _align("left", "top", wrap=True)
                elif ci == 7:
                    cell.font          = base_font
                    cell.alignment     = _align("center", "center", wrap=False)
                    cell.number_format = "#,##0"
                elif ci == 5:
                    khex       = PINNED_KIND_HEX if is_pin else KIND_HEX.get(item.kind, "374151")
                    cell.font  = Font(name="Arial", size=10, bold=True, color=khex)
                    cell.alignment = _align("center", "center", wrap=False)
                elif ci == 6:
                    pfont = Font(
                        name="Arial", size=10, bold=is_pin,
                        color=PINNED_KIND_HEX if is_pin else "374151",
                    )
                    cell.font      = pfont
                    cell.alignment = _align("center", "center", wrap=False)
                else:
                    col_align      = COLS[ci - 1][3]
                    cell.font      = base_font
                    cell.alignment = _align(col_align, "center", wrap=False)

            visible_lines = min(item.text.count("\n") + 1, XL_ROW_CONTENT_MAX_LINES)
            if visible_lines > 1:
                ws.row_dimensions[ri].height = XL_ROW_CONTENT_PER_LINE * visible_lines
            else:
                ws.row_dimensions[ri].height = XL_ROW_DATA

        last_letter = get_column_letter(N_COLS)
        ws.auto_filter.ref = f"A1:{last_letter}1"
        ws.freeze_panes    = "A2"

        for ci, (_lbl, col_min, col_max, _ha) in enumerate(COLS, 1):
            best = col_min
            for ri2 in range(1, min(len(items) + 2, 201)):
                cv = ws.cell(row=ri2, column=ci).value
                if cv is not None:
                    first_line_len = len(str(cv).split("\n")[0])
                    best = max(best, first_line_len + 2)
            ws.column_dimensions[get_column_letter(ci)].width = min(best, col_max)

        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth  = 1
        ws.page_setup.fitToHeight = 0
        ws.page_setup.orientation = "landscape"
        ws.page_setup.paperSize   = 9
        ws.print_title_rows       = "1:1"

        # ── Sheet 2 — Summary ─────────────────────────────────────────────
        ws2 = wb.create_sheet("Summary")
        ws2.column_dimensions["A"].width = 28
        ws2.column_dimensions["B"].width = 48
        ws2.column_dimensions["C"].width = 14
        ws2.column_dimensions["D"].width = 14

        TITLE_FONT = Font(name="Arial", bold=True,  size=15, color=XL["hdr_bg"])
        BOLD_FONT  = Font(name="Arial", bold=True,  size=11, color="111827")
        META_FONT  = Font(name="Arial", bold=False, size=11, color="374151")
        SMHDR_FONT = Font(name="Arial", bold=True,  size=11, color=XL["hdr_fg"])

        title_cell           = ws2.cell(
            row=1, column=1,
            value=_sanitize_for_xml(f"{APP_NAME}  —  Export Summary"),
        )
        title_cell.font      = TITLE_FONT
        title_cell.alignment = _align("left", "center", wrap=False)
        ws2.row_dimensions[1].height = 36
        ws2.merge_cells("A1:D1")

        export_path_str = _sanitize_for_xml(str(path.resolve()))
        meta_block: List[Tuple[str, object]] = [
            ("Export Date",  _sanitize_for_xml(datetime.now().strftime("%Y-%m-%d  %H:%M:%S"))),
            ("App Version",  _sanitize_for_xml(f"{APP_NAME}  v{APP_VERSION}")),
            ("Author",       _sanitize_for_xml(APP_AUTHOR)),
            ("Platform",     _sanitize_for_xml(f"{platform.system()} {platform.release()}")),
            ("Total Items",  len(items)),
            ("Pinned Items", sum(1 for i in items if i.pinned)),
            ("Export Path",  export_path_str),
        ]
        META_START = 3
        meta_fill  = _fill(XL["meta_bg"])

        for r_off, (label, value) in enumerate(meta_block, META_START):
            is_path_row = (label == "Export Path")
            lc = ws2.cell(row=r_off, column=1, value=_sanitize_for_xml(label))
            vc = ws2.cell(row=r_off, column=2, value=value)
            lc.font = BOLD_FONT
            vc.font = META_FONT

            if is_path_row:
                vc.alignment = _align("left", "top", wrap=True)
                col_b_chars   = 48
                path_chars    = len(export_path_str)
                wrapped_lines = max(1, -(-path_chars // col_b_chars))
                row_h         = float(min(max(22, wrapped_lines * 14), 80))
                ws2.row_dimensions[r_off].height = row_h
            else:
                vc.alignment = _align("left", "center", wrap=False)
                ws2.row_dimensions[r_off].height = 22

            lc.alignment = _align("left", "center", wrap=False)
            for cell in (lc, vc):
                cell.fill   = meta_fill
                cell.border = _border()

        STAT_ROW_START = META_START + len(meta_block) + 2

        stat_hdr_fill   = _fill(XL["stat_hdr"])
        stat_hdr_labels = ["Content Kind", "Label", "Count", "% of Total"]
        for ci, label in enumerate(stat_hdr_labels, 1):
            cell           = ws2.cell(row=STAT_ROW_START, column=ci,
                                      value=_sanitize_for_xml(label))
            cell.font      = SMHDR_FONT
            cell.fill      = stat_hdr_fill
            cell.alignment = _align("center", "center", wrap=False)
            cell.border    = _border()
        ws2.row_dimensions[STAT_ROW_START].height = 28

        kinds       = ("url", "email", "phone", "code", "text")
        kind_counts = {k: sum(1 for i in items if i.kind == k) for k in kinds}
        total_cnt   = max(len(items), 1)

        first_data_row: Optional[int] = None
        last_data_row : Optional[int] = None

        kind_label_full: Dict[str, str] = {
            "url"  : "URL / Hyperlink",
            "email": "Email Address",
            "phone": "Phone Number",
            "code" : "Code Snippet",
            "text" : "Plain Text",
        }

        for r_off, kind in enumerate(kinds, 1):
            cnt   = kind_counts[kind]
            row   = STAT_ROW_START + r_off
            rfill = _fill(XL["alt_row"] if r_off % 2 == 0 else XL["white"])
            khex  = KIND_HEX.get(kind, "374151")

            if first_data_row is None:
                first_data_row = row
            last_data_row = row

            cnt_cell               = ws2.cell(row=row, column=3, value=cnt)
            cnt_cell.font          = Font(name="Arial", size=10, bold=False, color=khex)
            cnt_cell.fill          = rfill
            cnt_cell.border        = _border()
            cnt_cell.alignment     = _align("center", "center", wrap=False)
            cnt_cell.number_format = "#,##0"

            pct_val                = cnt / total_cnt
            pct_cell               = ws2.cell(row=row, column=4, value=pct_val)
            pct_cell.font          = Font(name="Arial", size=10, bold=False, color=khex)
            pct_cell.fill          = rfill
            pct_cell.border        = _border()
            pct_cell.alignment     = _align("center", "center", wrap=False)
            pct_cell.number_format = "0.0%"

            for ci, val in enumerate(
                [_sanitize_for_xml(kind.upper()),
                 _sanitize_for_xml(kind_label_full.get(kind, kind))],
                1,
            ):
                cell           = ws2.cell(row=row, column=ci, value=val)
                cell.font      = Font(name="Arial", size=10, bold=(ci == 1), color=khex)
                cell.fill      = rfill
                cell.border    = _border()
                cell.alignment = _align("center", "center", wrap=False)

            ws2.row_dimensions[row].height = 22

        total_row  = STAT_ROW_START + len(kinds) + 1
        total_fill = _fill(XL["total"])

        if first_data_row is not None and last_data_row is not None:
            sum_formula = f"=SUM(C{first_data_row}:C{last_data_row})"
            pct_total   = 1.0
        else:
            sum_formula = 0
            pct_total   = 0.0

        total_row_vals = [
            (_sanitize_for_xml("TOTAL"),     None),
            (_sanitize_for_xml("All Kinds"), None),
            (sum_formula,                    "#,##0"),
            (pct_total,                      "0.0%"),
        ]
        for ci, (val, fmt) in enumerate(total_row_vals, 1):
            cell           = ws2.cell(row=total_row, column=ci, value=val)
            cell.font      = BOLD_FONT
            cell.fill      = total_fill
            cell.border    = _border()
            cell.alignment = _align("center", "center", wrap=False)
            if fmt:
                cell.number_format = fmt

        ws2.row_dimensions[total_row].height = 26
        ws2.freeze_panes = "A3"
        ws2.sheet_properties.pageSetUpPr.fitToPage = True
        ws2.page_setup.fitToWidth  = 1
        ws2.page_setup.fitToHeight = 0
        ws2.page_setup.orientation = "portrait"
        ws2.page_setup.paperSize   = 9

        tmp_path = path.with_suffix(".tmp")
        try:
            path.parent.mkdir(parents=True, exist_ok=True)
            wb.save(str(tmp_path))
            os.replace(tmp_path, path)
        except UnicodeEncodeError as exc:
            log.error("export_xlsx: UnicodeEncodeError during save: %s", exc)
            self._cleanup_tmp(tmp_path)
            raise
        # FIX-G: explicit catch for openpyxl IllegalCharacterError
        except Exception as exc:
            if _OpenpyxlIllegalChar and isinstance(exc, _OpenpyxlIllegalChar):
                log.error("export_xlsx: IllegalCharacterError — unexpected cell data: %s", exc)
            else:
                log.error("export_xlsx: %s during save: %s", type(exc).__name__, exc)
            self._cleanup_tmp(tmp_path)
            raise

        log.info("XLSX export: %d items → %s", len(items), path.resolve())
        return len(items)

    # ── Export: TXT ───────────────────────────────────────────────────────

    def export_txt(self, path: Path) -> int:
        """Export as structured UTF-8 (BOM) plain text. Atomic write."""
        items = self.all()
        WIDE = "═" * 80
        SLIM = "─" * 80
        now  = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        lines: List[str] = [
            WIDE,
            f"  {APP_NAME}",
            f"  Version   : v{APP_VERSION}",
            f"  Author    : {APP_AUTHOR}",
            f"  Exported  : {now}",
            f"  Platform  : {platform.system()} {platform.release()}",
            f"  Total     : {len(items):,} item{'s' if len(items) != 1 else ''}",
            f"  File      : {path.resolve()}",
            WIDE,
            "",
        ]

        for idx, item in enumerate(items, 1):
            pin_badge = "  ★ PINNED" if item.pinned else ""
            lines += [
                f"  Item #{idx:05d}   [{item.kind.upper()}]{pin_badge}",
                SLIM,
                f"  Timestamp  : {item.dt.strftime('%Y-%m-%d %H:%M:%S')}  ({item.human_time})",
                f"  Item ID    : {item.id}",
                f"  Characters : {len(item.text):,}",
                f"  Tags       : {', '.join(item.tags) if item.tags else '—'}",
                "",
                item.text,
                "",
                WIDE,
                "",
            ]

        tmp_path = path.with_suffix(".tmp")
        try:
            path.parent.mkdir(parents=True, exist_ok=True)
            tmp_path.write_text("\n".join(lines), encoding="utf-8-sig")
            os.replace(tmp_path, path)
        except OSError as exc:
            log.error("export_txt: save failed: %s", exc)
            self._cleanup_tmp(tmp_path)
            raise
        except Exception as exc:
            log.error("export_txt: %s during save: %s", type(exc).__name__, exc)
            self._cleanup_tmp(tmp_path)
            raise

        log.info("TXT export: %d items → %s", len(items), path.resolve())
        return len(items)

    # ── Export: JSON ──────────────────────────────────────────────────────

    def export_json(self, path: Path) -> int:
        """Export as structured JSON (UTF-8, 2-space indent). Atomic write."""
        items = self.all()
        doc = {
            "schema_version" : "1.0",
            "app"            : APP_NAME,
            "app_version"    : APP_VERSION,
            "author"         : APP_AUTHOR,
            "exported_at"    : datetime.now().isoformat(timespec="seconds"),
            "platform"       : f"{platform.system()} {platform.release()}",
            "statistics"     : {
                "total_items"  : len(items),
                "pinned_count" : sum(1 for i in items if i.pinned),
                "kind_counts"  : {
                    kind: sum(1 for i in items if i.kind == kind)
                    for kind in ("url", "email", "phone", "code", "text")
                },
            },
            "items" : [i.to_dict() for i in items],
        }
        tmp_path = path.with_suffix(".tmp")
        try:
            path.parent.mkdir(parents=True, exist_ok=True)
            tmp_path.write_text(
                json.dumps(doc, ensure_ascii=False, indent=2), encoding="utf-8"
            )
            os.replace(tmp_path, path)
        except UnicodeEncodeError as exc:
            log.error("export_json: UnicodeEncodeError during save: %s", exc)
            self._cleanup_tmp(tmp_path)
            raise
        except Exception as exc:
            log.error("export_json: %s during save: %s", type(exc).__name__, exc)
            self._cleanup_tmp(tmp_path)
            raise

        log.info("JSON export: %d items → %s", len(items), path.resolve())
        return len(items)

    # ── Lifecycle ─────────────────────────────────────────────────────────

    def stop(self) -> None:
        """Signal the flush thread to stop, wait for it, then do a final persist."""
        self._stop.set()
        self._dirty.set()
        self._flush_th.join(timeout=5)
        self._persist()

    # ── Private helpers ───────────────────────────────────────────────────

    def _mark_dirty(self) -> None:
        """Increment generation counter and signal flush thread. Call under _lock."""
        self._dirty_gen += 1
        self._dirty.set()

    @staticmethod
    def _cleanup_tmp(tmp_path: Path) -> None:
        try:
            if tmp_path.exists():
                os.unlink(str(tmp_path))
        except OSError:
            pass

    def _prune(self) -> None:
        """Remove excess unpinned items; updates _text_set. Call under _lock."""
        cap      = max(self._settings.max_items, MAX_ITEMS_MIN)
        pinned   = [i for i in self._items if i.pinned]
        unpinned = [i for i in self._items if not i.pinned]
        dropped  = unpinned[cap:]
        self._items = pinned + unpinned[:cap]
        for item in dropped:
            self._text_set.discard(item.text)

    def _load(self) -> None:
        """Load and deduplicate history from DB_FILE."""
        try:
            raw = json.loads(DB_FILE.read_text(encoding="utf-8"))
            if not isinstance(raw, list):
                raise ValueError("History file root must be a JSON array")

            loaded = [ClipItem.from_dict(d) for d in raw if isinstance(d, dict)]

            # Deduplicate — keep entry with newest timestamp per text
            seen: Dict[str, ClipItem] = {}
            for item in loaded:
                existing = seen.get(item.text)
                if existing is None or item.timestamp > existing.timestamp:
                    seen[item.text] = item

            deduped_ids = {id(v) for v in seen.values()}
            self._items = [i for i in loaded if id(i) in deduped_ids]

            self._text_set = {item.text for item in self._items}
            dupes = len(loaded) - len(self._items)
            if dupes:
                log.warning("Removed %d duplicate item(s) from history on load", dupes)
            log.info("Loaded %d clipboard items from history", len(self._items))
        except FileNotFoundError:
            log.info("No existing history — starting fresh")
        except (json.JSONDecodeError, ValueError, KeyError, TypeError) as exc:
            log.error("History file corrupt (%s) — backing up and starting fresh", exc)
            try:
                if DB_FILE.exists():
                    backup = DB_FILE.with_suffix(".json.bak")
                    DB_FILE.rename(backup)
                    log.info("Corrupt history backed up to %s", backup)
            except OSError:
                pass

    def _flush_loop(self) -> None:
        """
        Background thread: debounced disk write driven by generation counter.
        FIX-J: gen_before captured under _lock before event clear to eliminate
        race between clear() and gen read.
        """
        while not self._stop.is_set():
            self._dirty.wait(timeout=2.0)
            if self._stop.is_set():
                break

            # FIX-J: capture generation under lock BEFORE clearing the event
            with self._lock:
                gen_before = self._dirty_gen

            if gen_before == self._last_persisted_gen:
                self._dirty.clear()
                continue

            self._dirty.clear()
            self._persist()
            self._last_persisted_gen = gen_before

            # If new mutations occurred during persist, re-arm immediately
            with self._lock:
                if self._dirty_gen != gen_before:
                    self._dirty.set()

    def _persist(self) -> None:
        """Atomically write current items to DB_FILE."""
        tmp = DB_FILE.with_suffix(".tmp")
        try:
            with self._lock:
                payload = [i.to_dict() for i in self._items]
            APP_DATA_DIR.mkdir(parents=True, exist_ok=True)
            tmp.write_text(
                json.dumps(payload, ensure_ascii=False, separators=(",", ":")),
                encoding="utf-8",
            )
            os.replace(tmp, DB_FILE)
        except OSError as exc:
            log.error("DB persist failed: %s", exc)
            self._cleanup_tmp(tmp)


# ═════════════════════════════════════════════════════════════════════════════
# CLIPBOARD MONITOR
# ═════════════════════════════════════════════════════════════════════════════

class ClipboardMonitor:
    """
    Polls the system clipboard at POLL_INTERVAL_MS.
    Thread-safe; errors beyond first 3 logged every 10th occurrence.
    """

    def __init__(self, callback: Callable[[str], None]) -> None:
        self._cb        = callback
        self._lock      = threading.Lock()
        self._running   = False
        self._last      = ""
        self._last_ts   = 0.0
        self._err_count = 0
        self._thread    : Optional[threading.Thread] = None

    def start(self) -> None:
        with self._lock:
            if self._running:
                return
            try:
                self._last = _paste_with_timeout() or ""
            except Exception:
                self._last = ""
            self._running = True
            self._spawn_locked()
        log.info("Clipboard monitor started (poll interval: %d ms)", POLL_INTERVAL_MS)

    def stop(self) -> None:
        with self._lock:
            self._running = False
            thread = self._thread

        if thread and thread.is_alive():
            thread.join(timeout=2.0)

    @property
    def is_alive(self) -> bool:
        with self._lock:
            return bool(self._thread and self._thread.is_alive())

    def _spawn(self) -> None:
        """Public re-spawn entry point (used by watchdog). Thread-safe."""
        with self._lock:
            self._spawn_locked()

    def _spawn_locked(self) -> None:
        """Create and start the monitor thread. MUST be called with self._lock held."""
        if not self._running:
            return
        if self._thread and self._thread.is_alive():
            return
        self._thread = threading.Thread(
            target=self._loop, daemon=True, name="ClipMgr/Monitor"
        )
        self._thread.start()

    def _loop(self) -> None:
        interval = POLL_INTERVAL_MS / 1000.0
        while True:
            with self._lock:
                if not self._running:
                    break
            time.sleep(interval)

            try:
                current         = _paste_with_timeout()
                self._err_count = 0
                interval        = POLL_INTERVAL_MS / 1000.0
            except Exception as exc:
                self._err_count += 1
                interval = min(interval * 2, 30.0)
                if self._err_count <= 3 or self._err_count % 10 == 0:
                    log.warning(
                        "Clipboard read error #%d: %s  (next retry in %.1fs)",
                        self._err_count, exc, interval,
                    )
                continue

            now = time.time()
            if current and current != self._last and (now - self._last_ts) > DEBOUNCE_SEC:
                self._last    = current
                self._last_ts = now
                try:
                    self._cb(current)
                except Exception as exc:
                    log.error("Monitor callback raised an exception: %s", exc)

        log.debug("ClipboardMonitor loop exited")


# ═════════════════════════════════════════════════════════════════════════════
# GUI UTILITIES
# ═════════════════════════════════════════════════════════════════════════════

def _bind_recursive(widget, event: str, handler: Callable) -> None:
    """Attach an event handler to a widget and all its descendants."""
    widget.bind(event, handler, add="+")
    for child in widget.winfo_children():
        _bind_recursive(child, event, handler)


def _safe_after(widget, ms: int, fn: Callable) -> Optional[Any]:
    """Schedule fn via widget.after(); silently ignore if widget is destroyed."""
    try:
        return widget.after(ms, fn)
    except Exception:
        return None


def _safe_destroy(widget) -> None:
    """Destroy widget without raising if it is already gone."""
    try:
        widget.destroy()
    except Exception:
        pass


def _centre_window(window, width: int, height: int) -> None:
    window.update_idletasks()
    sw = window.winfo_screenwidth()
    sh = window.winfo_screenheight()
    x  = max(0, (sw - width)  // 2)
    y  = max(0, (sh - height) // 2)
    window.geometry(f"{width}x{height}+{x}+{y}")


def _centre_on_parent(child, parent) -> None:
    child.update_idletasks()
    pw = parent.winfo_width()
    ph = parent.winfo_height()
    px = parent.winfo_x()
    py = parent.winfo_y()
    cw = child.winfo_reqwidth()
    ch = child.winfo_reqheight()
    x  = max(0, px + (pw - cw) // 2)
    y  = max(0, py + (ph - ch) // 2)
    child.geometry(f"+{x}+{y}")


# ═════════════════════════════════════════════════════════════════════════════
# KIND DISPLAY MAPS
# ═════════════════════════════════════════════════════════════════════════════

KIND_BADGE: Dict[str, str] = {
    "url"  : "URL",
    "email": "MAIL",
    "phone": "PHONE",
    "code" : "CODE",
    "text" : "TEXT",
}
KIND_COLOR: Dict[str, str] = {
    "url"  : "#5B7FFF",
    "email": "#A78BFA",
    "phone": "#34D399",
    "code" : "#FBBF24",
    "text" : THEME["text_dim"],
}


# ═════════════════════════════════════════════════════════════════════════════
# CLIP CARD
# ═════════════════════════════════════════════════════════════════════════════

class ClipCard(ctk.CTkFrame):
    """
    A single clipboard history entry displayed as a card widget.
    Hover flicker fixed with _hover_depth counter (enter 0→1, leave 1→0).
    FIX-I: hover bindings use replace (not add="+") semantics via a per-card
    bind call; since each card is a fresh widget there is no prior binding to
    remove, but we deliberately do not use add="+" to be explicit.
    FIX-D: _after_ids typed as Set[Any] to handle platform-variant IDs.
    """

    def __init__(
        self,
        master,
        item      : ClipItem,
        on_copy   : Callable[[str], None],
        on_pin    : Callable[[str], None],
        on_delete : Callable[[str], None],
        font_size : int = 13,
        **kw,
    ) -> None:
        bg = THEME["card_pin"] if item.pinned else THEME["card"]
        super().__init__(master, fg_color=bg, corner_radius=10, **kw)
        self._item        = item
        self._on_copy     = on_copy
        self._on_pin      = on_pin
        self._on_delete   = on_delete
        self._alive       = True
        self._after_ids   : Set[Any] = set()     # FIX-D: Any, not str
        self._hover_depth : int = 0
        self._font_size   = font_size
        self._build()
        self._bind_hover()

    def _build(self) -> None:
        item = self._item
        fs   = self._font_size

        ctk.CTkLabel(
            self,
            text          = KIND_BADGE.get(item.kind, "TEXT"),
            font          = ctk.CTkFont(size=max(9, fs - 4), weight="bold"),
            text_color    = KIND_COLOR.get(item.kind, THEME["text_dim"]),
            fg_color      = THEME["border"],
            corner_radius = 6,
            width=50, height=22,
        ).grid(row=0, column=0, rowspan=2, padx=(10, 8), pady=10, sticky="n")

        ctk.CTkLabel(
            self,
            text       = item.preview,
            font       = ctk.CTkFont(family="Courier New", size=fs - 1),
            text_color = THEME["text"],
            anchor     = "w",
            wraplength = 480,
            justify    = "left",
        ).grid(row=0, column=1, sticky="sw", padx=(0, 4), pady=(10, 0))

        meta = ctk.CTkFrame(self, fg_color="transparent")
        meta.grid(row=1, column=1, sticky="nw", pady=(0, 8))

        ctk.CTkLabel(
            meta,
            text       = item.human_time,
            font       = ctk.CTkFont(size=max(9, fs - 3)),
            text_color = THEME["text_dim"],
        ).pack(side="left")

        ctk.CTkLabel(
            meta,
            text       = f"  ·  {len(item.text):,} characters",
            font       = ctk.CTkFont(size=max(9, fs - 3)),
            text_color = THEME["text_muted"],
        ).pack(side="left")

        if item.pinned:
            ctk.CTkLabel(
                meta,
                text       = "  ·  ★ Pinned",
                font       = ctk.CTkFont(size=max(9, fs - 3)),
                text_color = THEME["warn"],
            ).pack(side="left")

        btn_frame = ctk.CTkFrame(self, fg_color="transparent")
        btn_frame.grid(row=0, column=2, rowspan=2, padx=(0, 10), pady=8, sticky="e")

        self._copy_btn = ctk.CTkButton(
            btn_frame,
            text="Copy", width=58, height=28,
            font          = ctk.CTkFont(size=fs - 2, weight="bold"),
            fg_color      = THEME["accent"],
            hover_color   = "#4A6BE0",
            corner_radius = 7,
            command       = self._copy,
        )
        self._copy_btn.pack(pady=(0, 4))

        ctk.CTkButton(
            btn_frame,
            text          = "Unpin" if item.pinned else "Pin",
            width=58, height=28,
            font          = ctk.CTkFont(size=fs - 2),
            fg_color      = THEME["warn"] if item.pinned else THEME["border"],
            hover_color   = THEME["card_hover"],
            corner_radius = 7,
            command       = self._pin,
        ).pack(pady=(0, 4))

        ctk.CTkButton(
            btn_frame,
            text          = "Delete",
            width=58, height=28,
            font          = ctk.CTkFont(size=fs - 2),
            fg_color      = THEME["border"],
            hover_color   = THEME["danger"],
            corner_radius = 7,
            command       = self._delete,
        ).pack()

        self.columnconfigure(1, weight=1)

    def _bind_hover(self) -> None:
        """
        Depth-counter hover: only outer-most enter/leave toggles colour.
        FIX-I: do NOT use add="+" — avoid binding accumulation.
        """
        bg_norm  = THEME["card_pin"] if self._item.pinned else THEME["card"]
        bg_hover = THEME["card_hover"]

        def _on_enter(_e):
            self._hover_depth += 1
            if self._hover_depth == 1:
                try:
                    self.configure(fg_color=bg_hover)
                except Exception:
                    pass

        def _on_leave(_e):
            self._hover_depth = max(0, self._hover_depth - 1)
            if self._hover_depth == 0:
                try:
                    self.configure(fg_color=bg_norm)
                except Exception:
                    pass

        # FIX-I: bind without add="+" so bindings don't accumulate
        def _bind_no_add(widget):
            widget.bind("<Enter>", _on_enter)
            widget.bind("<Leave>", _on_leave)
            for child in widget.winfo_children():
                _bind_no_add(child)

        _bind_no_add(self)

    def destroy(self) -> None:
        self._alive = False
        for aid in list(self._after_ids):
            try:
                self.after_cancel(aid)
            except Exception:
                pass
        self._after_ids.clear()
        super().destroy()

    def _copy(self) -> None:
        self._on_copy(self._item.id)
        try:
            if self._copy_btn.winfo_exists():
                self._copy_btn.configure(text="✔ Copied", fg_color=THEME["success"])
        except Exception:
            return

        orig = THEME["accent"]

        def _restore():
            if self._alive:
                try:
                    if self._copy_btn.winfo_exists():
                        self._copy_btn.configure(text="Copy", fg_color=orig)
                except Exception:
                    pass

        aid = _safe_after(self, 1_400, _restore)
        if aid is not None:
            self._after_ids.add(aid)

    def _pin(self)    -> None: self._on_pin(self._item.id)
    def _delete(self) -> None: self._on_delete(self._item.id)


# ═════════════════════════════════════════════════════════════════════════════
# SCROLLABLE CLIP LIST
# ═════════════════════════════════════════════════════════════════════════════

class ScrollableClipList(ctk.CTkScrollableFrame):
    """
    Scrollable container that renders ClipCard widgets.
    Deferred render lambda guards winfo_exists() inside the lambda body.
    FIX-K: scrolls to top after render to reveal newest items.
    """

    def __init__(self, master, font_size: int = 13, **kw) -> None:
        super().__init__(
            master,
            fg_color                     = THEME["bg"],
            scrollbar_button_color       = THEME["scrollbar"],
            scrollbar_button_hover_color = THEME["accent"],
            **kw,
        )
        self._cards     : List[Any] = []
        self._rendering : bool = False
        self._pending   : Optional[Tuple] = None
        self._font_size : int = font_size

    def set_font_size(self, font_size: int) -> None:
        self._font_size = font_size

    def render(
        self,
        items     : List[ClipItem],
        on_copy   : Callable,
        on_pin    : Callable,
        on_delete : Callable,
    ) -> None:
        if self._rendering:
            self._pending = (items, on_copy, on_pin, on_delete)
            return

        self._rendering = True
        try:
            for widget in self._cards:
                try:
                    widget.destroy()
                except Exception:
                    pass
            self._cards.clear()

            if not items:
                placeholder = ctk.CTkLabel(
                    self,
                    text       = "No clipboard entries yet — copy something to get started.",
                    text_color = THEME["text_muted"],
                    font       = ctk.CTkFont(size=self._font_size + 1),
                )
                placeholder.pack(pady=80)
                self._cards.append(placeholder)
                return

            for item in items[:VIRTUAL_PAGE]:
                card = ClipCard(
                    self, item,
                    on_copy=on_copy, on_pin=on_pin, on_delete=on_delete,
                    font_size=self._font_size,
                )
                card.pack(fill="x", padx=6, pady=3)
                self._cards.append(card)

            if len(items) > VIRTUAL_PAGE:
                overflow = ctk.CTkLabel(
                    self,
                    text       = (
                        f"Showing {VIRTUAL_PAGE} of {len(items)} items.  "
                        "Use the search bar to narrow results."
                    ),
                    text_color = THEME["text_muted"],
                    font       = ctk.CTkFont(size=self._font_size - 2),
                )
                overflow.pack(pady=(4, 16))
                self._cards.append(overflow)

            # FIX-K: scroll to top so newest items are visible
            try:
                self._parent_canvas.yview_moveto(0)
            except Exception:
                pass

        finally:
            self._rendering = False
            if self._pending is not None:
                args          = self._pending
                self._pending = None
                try:
                    if self.winfo_exists():
                        self.after(
                            0,
                            lambda a=args: self.render(*a) if self.winfo_exists() else None
                        )
                except Exception:
                    pass


# ═════════════════════════════════════════════════════════════════════════════
# SETTINGS DIALOG
# ═════════════════════════════════════════════════════════════════════════════

class SettingsDialog(ctk.CTkToplevel):
    """Modal settings dialog, centred over the main window."""

    _WIN_W = 460
    _WIN_H = 470

    def __init__(self, master, settings: Settings, on_save: Callable[[], None]) -> None:
        super().__init__(master)
        self.title("Settings")
        self.resizable(False, False)
        self.configure(fg_color=THEME["panel"])
        self.transient(master)

        # FIX-E: guard grab_set() so a destroyed parent doesn't crash us
        try:
            self.grab_set()
        except Exception:
            pass

        self._master   = master
        self._settings = settings
        self._on_save  = on_save
        self._build()
        _centre_on_parent(self, master)

        self.bind("<Return>", lambda _e: self._save())
        self.bind("<Escape>", lambda _e: self._safe_close())
        # FIX-E: always release grab before destroying
        self.protocol("WM_DELETE_WINDOW", self._safe_close)

    def _safe_close(self) -> None:
        """Release grab safely, then destroy."""
        try:
            self.grab_release()
        except Exception:
            pass
        try:
            self.destroy()
        except Exception:
            pass

    def _build(self) -> None:
        s   = self._settings
        PAD = {"padx": 24, "pady": 7}

        ctk.CTkLabel(
            self, text="Settings",
            font       = ctk.CTkFont(size=18, weight="bold"),
            text_color = THEME["accent"],
        ).pack(anchor="w", padx=24, pady=(22, 10))

        row = ctk.CTkFrame(self, fg_color=THEME["card"], corner_radius=10)
        row.pack(fill="x", **PAD)
        ctk.CTkLabel(
            row,
            text       = f"Maximum history items  ({MAX_ITEMS_MIN} – {MAX_ITEMS_MAX})",
            text_color = THEME["text"],
        ).pack(side="left", padx=14, pady=12)
        self._max_var = ctk.StringVar(value=str(s.max_items))
        ctk.CTkEntry(
            row, textvariable=self._max_var, width=82,
            fg_color=THEME["bg"], border_color=THEME["border"],
        ).pack(side="right", padx=14)

        self._dedup_var = ctk.BooleanVar(value=s.dedup)
        self._trim_var  = ctk.BooleanVar(value=s.trim_whitespace)
        self._minim_var = ctk.BooleanVar(value=s.start_minimized)

        toggles = [
            ("Deduplicate entries",              self._dedup_var),
            ("Trim leading/trailing whitespace", self._trim_var),
            ("Start minimised to taskbar",       self._minim_var),
        ]
        for label, var in toggles:
            r = ctk.CTkFrame(self, fg_color=THEME["card"], corner_radius=10)
            r.pack(fill="x", **PAD)
            ctk.CTkLabel(r, text=label, text_color=THEME["text"]).pack(
                side="left", padx=14, pady=10
            )
            ctk.CTkSwitch(
                r, text="", variable=var,
                progress_color=THEME["accent"],
                button_color=THEME["text"],
            ).pack(side="right", padx=14)

        row2 = ctk.CTkFrame(self, fg_color=THEME["card"], corner_radius=10)
        row2.pack(fill="x", **PAD)
        ctk.CTkLabel(
            row2,
            text       = f"Interface font size  ({FONT_SIZE_MIN} – {FONT_SIZE_MAX})",
            text_color = THEME["text"],
        ).pack(side="left", padx=14, pady=12)
        self._font_var = ctk.StringVar(value=str(s.font_size))
        ctk.CTkEntry(
            row2, textvariable=self._font_var, width=62,
            fg_color=THEME["bg"], border_color=THEME["border"],
        ).pack(side="right", padx=14)

        btns = ctk.CTkFrame(self, fg_color="transparent")
        btns.pack(fill="x", padx=24, pady=18)
        ctk.CTkButton(
            btns, text="Cancel", width=100,
            fg_color    = THEME["border"],
            hover_color = THEME["card_hover"],
            command     = self._safe_close,
        ).pack(side="left")
        ctk.CTkButton(
            btns, text="Save Settings", width=130,
            fg_color    = THEME["accent"],
            hover_color = "#4A6BE0",
            command     = self._save,
        ).pack(side="right")

    def _save(self) -> None:
        raw_max = self._max_var.get().strip()
        try:
            max_val = int(raw_max)
        except ValueError:
            _toast(
                self._master,
                f"Max items must be a whole number between {MAX_ITEMS_MIN} and {MAX_ITEMS_MAX}.",
                kind="warn",
            )
            return
        if not (MAX_ITEMS_MIN <= max_val <= MAX_ITEMS_MAX):
            _toast(
                self._master,
                f"Max items must be between {MAX_ITEMS_MIN} and {MAX_ITEMS_MAX}.",
                kind="warn",
            )
            return

        raw_font = self._font_var.get().strip()
        try:
            font_val = int(raw_font)
        except ValueError:
            _toast(
                self._master,
                f"Font size must be a whole number between {FONT_SIZE_MIN} and {FONT_SIZE_MAX}.",
                kind="warn",
            )
            return
        if not (FONT_SIZE_MIN <= font_val <= FONT_SIZE_MAX):
            _toast(
                self._master,
                f"Font size must be between {FONT_SIZE_MIN} and {FONT_SIZE_MAX}.",
                kind="warn",
            )
            return

        self._settings.max_items       = max_val
        self._settings.font_size       = font_val
        self._settings.dedup           = self._dedup_var.get()
        self._settings.trim_whitespace = self._trim_var.get()
        self._settings.start_minimized = self._minim_var.get()
        self._on_save()
        self._safe_close()


# ═════════════════════════════════════════════════════════════════════════════
# TOAST NOTIFICATION
# ═════════════════════════════════════════════════════════════════════════════

def _toast(
    master,
    message  : str,
    kind     : str = "success",
    duration : int = 2_800,
) -> None:
    """Display a transient notification near the bottom-right corner."""
    color = {
        "success": THEME["success"],
        "warn"   : THEME["warn"],
        "danger" : THEME["danger"],
    }.get(kind, THEME["accent"])

    try:
        if not master.winfo_exists():
            return

        toast = ctk.CTkToplevel(master)
        toast.overrideredirect(True)
        toast.attributes("-topmost", True)
        toast.configure(fg_color=THEME["panel"])

        ctk.CTkLabel(
            toast,
            text       = message,
            text_color = color,
            font       = ctk.CTkFont(size=13, weight="bold"),
            padx=22, pady=12,
        ).pack()

        toast.update_idletasks()
        tw  = toast.winfo_reqwidth()
        th  = toast.winfo_reqheight()
        sw  = master.winfo_screenwidth()
        sh  = master.winfo_screenheight()
        mw  = master.winfo_width()
        mh  = master.winfo_height()
        mx  = master.winfo_rootx()
        my  = master.winfo_rooty()

        raw_x = (mx + mw - tw - 14) if mw > 1 else (sw - tw - 20)
        raw_y = (my + mh - th - 14) if mh > 1 else (sh - th - 60)
        x = max(0, min(raw_x, sw - tw - 4))
        y = max(0, min(raw_y, sh - th - 48))
        toast.geometry(f"+{x}+{y}")

        def _dismiss() -> None:
            try:
                if toast.winfo_exists():
                    toast.destroy()
            except Exception:
                pass

        try:
            master.after(duration, _dismiss)
        except Exception:
            _dismiss()

    except Exception as exc:
        log.warning("Toast notification failed: %s", exc)


# ═════════════════════════════════════════════════════════════════════════════
# MAIN APPLICATION WINDOW
# ═════════════════════════════════════════════════════════════════════════════

class ClipboardManagerApp(ctk.CTk):
    """
    Main application window for Clipboard History Manager v1.0.0.

    Shutdown order:
      1.  _running = False
      1b. withdraw() immediately
      2.  Cancel named recurring after() IDs
      3.  Cancel remaining one-shot after() IDs
      4.  Cancel search debounce timer
      5.  Stop hotkey listener
      6.  Stop clipboard monitor
      6b. Drain _ui_queue into store
      7.  store.stop() on background thread
      8.  Poll main thread until store done, then destroy (FIX-C)

    FIX-C replaces the unsafe self.after() call from a non-Tk thread with a
    main-thread polling pattern using a threading.Event.
    """

    _WIN_W = 900
    _WIN_H = 760

    def __init__(self) -> None:
        ctk.set_appearance_mode("dark")
        ctk.set_default_color_theme("blue")
        super().__init__()

        self._settings        = Settings.load()
        self._store           = HistoryStore(self._settings)
        self._monitor         = ClipboardMonitor(self._on_new_clip)
        self._ui_queue        : queue.SimpleQueue = queue.SimpleQueue()

        self._filter          : str            = ""
        self._tab             : str            = "all"
        self._search_timer    : Optional[Any]  = None
        self._hotkey_listener : Optional[Any]  = None

        self._running         : bool           = True
        self._queue_poll_id   : Optional[Any]  = None
        self._watchdog_id     : Optional[Any]  = None
        self._after_ids       : Set[Any]       = set()  # FIX-D

        # FIX-C: event set by shutdown thread; main thread polls to destroy
        self._store_done_event : threading.Event = threading.Event()

        self._configure_window()
        self._build_ui()

        self._setup_global_hotkey()
        self._monitor.start()
        self._schedule_queue_poll()
        self._refresh_list()
        self._schedule_watchdog()

        self.protocol("WM_DELETE_WINDOW", self._on_close)
        self.bind("<Control-f>", lambda _e: (self._search_entry.focus_set(), "break"))
        self.bind("<Escape>",    lambda _e: self._clear_search())

        if self._settings.start_minimized:
            self.after(100, self.iconify)

        log.info("%s v%s initialised successfully", APP_NAME, APP_VERSION)

    # ── Window configuration ──────────────────────────────────────────────

    def _configure_window(self) -> None:
        self.title(f"📋  {APP_NAME}  —  v{APP_VERSION}")
        self.minsize(720, 520)
        self.configure(fg_color=THEME["bg"])
        _centre_window(self, self._WIN_W, self._WIN_H)

    # ── Font size helpers ─────────────────────────────────────────────────

    def _fs(self, delta: int = 0) -> int:
        return max(FONT_SIZE_MIN, min(FONT_SIZE_MAX, self._settings.font_size + delta))

    def _rebuild_fonts(self) -> None:
        self._clip_list.set_font_size(self._fs())
        self._refresh_list()

    # ── Tracked after() helpers ───────────────────────────────────────────

    def _safe_after_tracked(self, ms: int, fn: Callable) -> Optional[Any]:
        """Schedule fn and track the ID for cancellation on shutdown."""
        aid_holder: List[Optional[Any]] = [None]

        def _wrapped():
            aid = aid_holder[0]
            if aid is not None:
                self._after_ids.discard(aid)
            fn()

        try:
            aid = self.after(ms, _wrapped)
            aid_holder[0] = aid
            self._after_ids.add(aid)
            return aid
        except Exception:
            return None

    # ═════════════════════════════════════════════════════════════════════
    # UI Construction
    # ═════════════════════════════════════════════════════════════════════

    def _build_ui(self) -> None:
        self._build_header()
        self._build_toolbar()
        self._build_tabs()
        self._build_list()
        self._build_statusbar()

    def _build_header(self) -> None:
        hdr = ctk.CTkFrame(self, fg_color=THEME["panel"], corner_radius=0, height=66)
        hdr.pack(fill="x")
        hdr.pack_propagate(False)

        ctk.CTkLabel(
            hdr,
            text       = "📋  Clipboard History Manager",
            font       = ctk.CTkFont(size=19, weight="bold"),
            text_color = THEME["text"],
        ).pack(side="left", padx=22, pady=18)

        btn_container = ctk.CTkFrame(hdr, fg_color="transparent")
        btn_container.pack(side="right", padx=18)

        ctk.CTkButton(
            btn_container, text="↑ Export", width=96, height=36,
            font          = ctk.CTkFont(size=12, weight="bold"),
            fg_color      = THEME["card"],
            hover_color   = THEME["accent"],
            corner_radius = 8,
            command       = self._show_export_menu,
        ).pack(side="right", padx=4)

        ctk.CTkButton(
            btn_container, text="⚙  Settings", width=100, height=36,
            font          = ctk.CTkFont(size=12),
            fg_color      = THEME["border"],
            hover_color   = THEME["accent"],
            corner_radius = 8,
            command       = self._open_settings,
        ).pack(side="right", padx=4)

    def _build_toolbar(self) -> None:
        bar = ctk.CTkFrame(self, fg_color=THEME["panel"], corner_radius=0, height=58)
        bar.pack(fill="x")
        bar.pack_propagate(False)

        search_wrap = ctk.CTkFrame(bar, fg_color=THEME["card"], corner_radius=10)
        search_wrap.pack(side="left", padx=18, pady=11, fill="x", expand=True)

        ctk.CTkLabel(
            search_wrap, text="🔍",
            font=ctk.CTkFont(size=14), text_color=THEME["text_dim"],
        ).pack(side="left", padx=(14, 4))

        self._search_var = ctk.StringVar()
        self._search_var.trace_add("write", self._on_search_change)
        self._search_entry = ctk.CTkEntry(
            search_wrap,
            textvariable     = self._search_var,
            placeholder_text = "Search clipboard history…   (Ctrl+F to focus)",
            fg_color         = "transparent",
            border_width     = 0,
            font             = ctk.CTkFont(size=self._fs()),
            text_color       = THEME["text"],
        )
        self._search_entry.pack(side="left", fill="x", expand=True, pady=7)

        ctk.CTkButton(
            search_wrap, text="✕", width=28, height=28,
            font          = ctk.CTkFont(size=11),
            fg_color      = "transparent",
            hover_color   = THEME["border"],
            corner_radius = 6,
            command       = self._clear_search,
        ).pack(side="right", padx=5)

        ctk.CTkButton(
            bar, text="🗑  Clear History", width=118, height=36,
            font          = ctk.CTkFont(size=12),
            fg_color      = THEME["card"],
            hover_color   = THEME["danger"],
            corner_radius = 8,
            command       = self._confirm_clear,
        ).pack(side="right", padx=18, pady=11)

    def _build_tabs(self) -> None:
        tab_bar = ctk.CTkFrame(self, fg_color=THEME["panel"], corner_radius=0, height=46)
        tab_bar.pack(fill="x")
        tab_bar.pack_propagate(False)

        self._tab_btns: Dict[str, ctk.CTkButton] = {}
        for key, label in [("all", "All Items"), ("pinned", "📌  Pinned")]:
            btn = ctk.CTkButton(
                tab_bar, text=label, width=136, height=32,
                font          = ctk.CTkFont(size=12, weight="bold"),
                fg_color      = THEME["accent"] if key == self._tab else THEME["card"],
                hover_color   = THEME["card_hover"],
                corner_radius = 8,
                command       = lambda k=key: self._switch_tab(k),
            )
            btn.pack(side="left", padx=(18 if key == "all" else 4, 0), pady=7)
            self._tab_btns[key] = btn

        self._count_label = ctk.CTkLabel(
            tab_bar, text="",
            font=ctk.CTkFont(size=11), text_color=THEME["text_dim"],
        )
        self._count_label.pack(side="right", padx=18)

    def _build_list(self) -> None:
        self._clip_list = ScrollableClipList(self, font_size=self._fs())
        self._clip_list.pack(fill="both", expand=True, padx=8, pady=(4, 0))

    def _build_statusbar(self) -> None:
        sb = ctk.CTkFrame(self, fg_color=THEME["panel"], corner_radius=0, height=28)
        sb.pack(fill="x", side="bottom")
        sb.pack_propagate(False)

        self._status_label = ctk.CTkLabel(
            sb,
            text       = "●  Monitoring clipboard",
            font       = ctk.CTkFont(size=10),
            text_color = THEME["success"],
        )
        self._status_label.pack(side="left", padx=16)

        hints: List[str] = [f"Data folder: {APP_DATA_DIR}"]
        if PYNPUT_OK:
            hints.append("Ctrl+Shift+V — show window")
        if not OPENPYXL_OK:
            hints.append("pip install openpyxl — enable XLSX export")
        ctk.CTkLabel(
            sb,
            text       = "   |   ".join(hints),
            font       = ctk.CTkFont(size=10),
            text_color = THEME["text_muted"],
        ).pack(side="right", padx=16)

    # ═════════════════════════════════════════════════════════════════════
    # Watchdog
    # ═════════════════════════════════════════════════════════════════════

    def _schedule_watchdog(self) -> None:
        if not self._running:
            return
        try:
            self._watchdog_id = self.after(WATCHDOG_MS, self._watchdog)
        except Exception:
            pass

    def _watchdog(self) -> None:
        if not self._running:
            return

        if self._monitor._running and not self._monitor.is_alive:
            log.warning("Clipboard monitor thread died unexpectedly — restarting")
            self._monitor._spawn()
            try:
                self._status_label.configure(
                    text="⚠  Monitor restarted", text_color=THEME["warn"]
                )
            except Exception:
                pass

            def _restore_status():
                try:
                    self._status_label.configure(
                        text="●  Monitoring clipboard", text_color=THEME["success"]
                    )
                except Exception:
                    pass

            self._safe_after_tracked(4_000, _restore_status)

        self._schedule_watchdog()

    # ═════════════════════════════════════════════════════════════════════
    # Global hotkey  (Ctrl+Shift+V)
    # ═════════════════════════════════════════════════════════════════════

    def _setup_global_hotkey(self) -> None:
        if not PYNPUT_OK:
            return
        listener: Optional[Any] = None
        try:
            hotkey = _pynput_keyboard.HotKey(
                _pynput_keyboard.HotKey.parse("<ctrl>+<shift>+v"),
                self._hotkey_triggered,
            )

            def _on_press(key):
                try:
                    hotkey.press(
                        _pynput_keyboard.KeyCode.from_vk(key.vk)
                        if hasattr(key, "vk") else key
                    )
                except Exception:
                    pass

            def _on_release(key):
                try:
                    hotkey.release(
                        _pynput_keyboard.KeyCode.from_vk(key.vk)
                        if hasattr(key, "vk") else key
                    )
                except Exception:
                    pass

            listener = _pynput_keyboard.Listener(
                on_press=_on_press, on_release=_on_release
            )
            listener.daemon = True
            listener.start()
            self._hotkey_listener = listener
            log.info("Global hotkey registered: Ctrl+Shift+V")
        except Exception as exc:
            log.warning("Could not register global hotkey: %s", exc)
            if listener is not None:
                try:
                    listener.stop()
                except Exception:
                    pass

    def _hotkey_triggered(self) -> None:
        self.after(0, self._bring_to_front)

    def _bring_to_front(self) -> None:
        self.deiconify()
        self.lift()
        self.focus_force()

    # ═════════════════════════════════════════════════════════════════════
    # Core data flow
    # ═════════════════════════════════════════════════════════════════════

    def _on_new_clip(self, text: str) -> None:
        """Called from monitor thread — enqueue for main-thread processing."""
        self._ui_queue.put(("add", text))

    def _schedule_queue_poll(self) -> None:
        if not self._running:
            return
        try:
            self._queue_poll_id = self.after(150, self._process_ui_queue)
        except Exception:
            pass

    def _process_ui_queue(self) -> None:
        """Drain the inter-thread queue on the Tk main thread."""
        if not self._running:
            return

        refreshed = False
        drained   = 0
        while drained < QUEUE_DRAIN_MAX:
            try:
                kind, payload = self._ui_queue.get_nowait()
                drained += 1
                try:
                    if kind == "add" and self._store.add(payload):
                        refreshed = True
                except Exception as exc:
                    log.error("UI queue item processing error: %s", exc)
            except queue.Empty:
                break
            except Exception as exc:
                log.error("UI queue drain error: %s", exc)
                break

        if refreshed:
            self._refresh_list()

        self._schedule_queue_poll()

    def _refresh_list(self) -> None:
        """Re-query the store and render the visible card list."""
        query = self._filter.strip()
        items = self._store.search(query) if query else self._store.all()

        if self._tab == "pinned":
            items = [i for i in items if i.pinned]

        self._clip_list.render(
            items,
            on_copy   = self._copy_item,
            on_pin    = self._pin_item,
            on_delete = self._delete_item,
        )

        all_items  = self._store.all()
        total      = len(all_items)
        pinned_cnt = sum(1 for i in all_items if i.pinned)
        self._count_label.configure(
            text=f"{len(items)} shown  ·  {total} total  ·  {pinned_cnt} pinned"
        )

    # ── Item actions ──────────────────────────────────────────────────────

    def _copy_item(self, item_id: str) -> None:
        """Copy runs in a background thread via _copy_with_timeout()."""
        item = self._store.get(item_id)
        if not item:
            return
        try:
            _copy_with_timeout(item.text)
            self._status_label.configure(
                text       = f"●  Copied: {item.preview[:55]}",
                text_color = THEME["success"],
            )
            self._safe_after_tracked(3_000, lambda: self._status_label.configure(
                text="●  Monitoring clipboard", text_color=THEME["success"]
            ))
        except Exception as exc:
            log.error("Failed to copy item to clipboard: %s", exc)
            _toast(self, f"Copy failed: {exc}", kind="danger")

    def _pin_item(self, item_id: str) -> None:
        self._store.pin(item_id)
        self._refresh_list()

    def _delete_item(self, item_id: str) -> None:
        self._store.delete(item_id)
        self._refresh_list()

    # ── Search ────────────────────────────────────────────────────────────

    def _on_search_change(self, *_) -> None:
        if self._search_timer is not None:
            try:
                self.after_cancel(self._search_timer)
            except Exception:
                pass
        self._search_timer = self.after(SEARCH_DEBOUNCE, self._apply_search)

    def _apply_search(self) -> None:
        self._search_timer = None
        self._filter = self._search_var.get()
        self._refresh_list()

    def _clear_search(self) -> None:
        self._search_var.set("")
        self._filter = ""
        self._refresh_list()

    # ── Tabs ──────────────────────────────────────────────────────────────

    def _switch_tab(self, tab: str) -> None:
        self._tab = tab
        for key, btn in self._tab_btns.items():
            btn.configure(fg_color=THEME["accent"] if key == tab else THEME["card"])
        self._refresh_list()

    # ── Clear history ─────────────────────────────────────────────────────

    def _confirm_clear(self) -> None:
        dlg = ctk.CTkToplevel(self)
        dlg.title("Confirm Clear History")
        dlg.resizable(False, False)
        dlg.configure(fg_color=THEME["panel"])
        dlg.transient(self)
        try:
            dlg.grab_set()
        except Exception:
            pass

        ctk.CTkLabel(
            dlg,
            text       = "Clear clipboard history?",
            font       = ctk.CTkFont(size=15, weight="bold"),
            text_color = THEME["text"],
        ).pack(pady=(26, 4))
        ctk.CTkLabel(
            dlg,
            text       = "All unpinned items will be permanently removed.\nPinned items are kept.",
            text_color = THEME["text_dim"],
            font       = ctk.CTkFont(size=12),
            justify    = "center",
        ).pack(pady=(0, 22))

        btns = ctk.CTkFrame(dlg, fg_color="transparent")
        btns.pack(pady=(0, 22))

        def _safe_close_dlg():
            try:
                dlg.grab_release()
            except Exception:
                pass
            try:
                dlg.destroy()
            except Exception:
                pass

        def do_clear():
            removed = self._store.clear_unpinned()
            _safe_close_dlg()
            self._refresh_list()
            _toast(self, f"Cleared {removed} item{'s' if removed != 1 else ''}", kind="warn")

        ctk.CTkButton(
            btns, text="Cancel", width=110,
            fg_color    = THEME["border"],
            hover_color = THEME["card_hover"],
            command     = _safe_close_dlg,
        ).pack(side="left", padx=8)

        ctk.CTkButton(
            btns, text="Clear History", width=130,
            fg_color    = THEME["danger"],
            hover_color = "#C0392B",
            command     = do_clear,
        ).pack(side="left", padx=8)

        dlg.bind("<Return>", lambda _e: do_clear())
        dlg.bind("<Escape>", lambda _e: _safe_close_dlg())
        dlg.protocol("WM_DELETE_WINDOW", _safe_close_dlg)
        _centre_on_parent(dlg, self)

    # ── Export ────────────────────────────────────────────────────────────

    def _show_export_menu(self) -> None:
        """
        FIX-A: Uses a named handler stored in a local variable for unbinding,
        instead of unbind_all() which would remove ALL Button-1 bindings
        application-wide.
        FIX-F: _arm_close guards winfo_exists() before binding; timer reduced
        to 80 ms to minimise the destroyed-widget window.
        """
        menu = ctk.CTkToplevel(self)
        menu.overrideredirect(True)
        menu.configure(fg_color=THEME["card"])
        menu.attributes("-topmost", True)

        XLSX_LABEL = (
            "📗  Export as Excel (.xlsx)  ★ Primary"
            if OPENPYXL_OK
            else "📗  Excel (.xlsx)  —  pip install openpyxl"
        )
        options: List[Tuple[str, Callable, bool]] = [
            (XLSX_LABEL,                               self._export_xlsx, OPENPYXL_OK),
            ("📄  Export as Plain Text (.txt)  Backup", self._export_txt,  True),
            ("🔧  Export as JSON (.json)  Power User",  self._export_json, True),
        ]

        for label, fn, enabled in options:
            btn = ctk.CTkButton(
                menu, text=label, width=300, height=42,
                font        = ctk.CTkFont(size=12),
                fg_color    = "transparent",
                hover_color = THEME["accent"] if enabled else THEME["border"],
                text_color  = THEME["text"] if enabled else THEME["text_muted"],
                anchor      = "w",
                state       = "normal" if enabled else "disabled",
                command     = lambda f=fn: (menu.destroy(), f()),
            )
            btn.pack(fill="x", padx=4, pady=2)

        self.update_idletasks()
        mx = self.winfo_x() + self.winfo_width() - 330
        my = self.winfo_y() + 70
        menu.geometry(f"316x{len(options) * 50 + 8}+{mx}+{my}")

        # FIX-A: keep a reference to the handler so we can unbind it
        # specifically, without using unbind_all() which destroys all bindings.
        _outside_click_handler_id: List[Optional[str]] = [None]

        def _check_outside_click(event):
            try:
                if not menu.winfo_exists():
                    return
                w = event.widget
                while w is not None:
                    if w is menu:
                        return
                    try:
                        w = w.master
                    except AttributeError:
                        break
                _safe_destroy(menu)
            except Exception:
                _safe_destroy(menu)

        def _unbind_outside_handler():
            """Unbind only our specific handler, not all Button-1 bindings."""
            hid = _outside_click_handler_id[0]
            if hid is not None:
                try:
                    self.unbind("<Button-1>", hid)
                except Exception:
                    pass
                _outside_click_handler_id[0] = None

        def _arm_close():
            # FIX-F: guard winfo_exists() before doing anything
            try:
                if not menu.winfo_exists():
                    return
                hid = self.bind("<Button-1>", _check_outside_click, add="+")
                _outside_click_handler_id[0] = hid
                menu.bind(
                    "<Destroy>",
                    lambda _e: _unbind_outside_handler(),
                    add="+",
                )
            except Exception:
                pass

        # FIX-F: 80 ms (reduced from 130 ms) to narrow the destroyed-widget window
        menu.after(80, _arm_close)

    def _export_xlsx(self) -> None: self._do_export("xlsx")
    def _export_txt(self)  -> None: self._do_export("txt")
    def _export_json(self) -> None: self._do_export("json")

    def _do_export(self, fmt: str) -> None:
        from tkinter import filedialog

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        defaults  = {
            "xlsx": (f"clipboard_history_{timestamp}.xlsx", [("Excel Workbook", "*.xlsx")]),
            "txt" : (f"clipboard_history_{timestamp}.txt",  [("Plain Text File", "*.txt")]),
            "json": (f"clipboard_history_{timestamp}.json", [("JSON File",       "*.json")]),
        }
        default_name, ftypes = defaults.get(fmt, (f"clipboard_history_{timestamp}.{fmt}", []))

        path_str = filedialog.asksaveasfilename(
            defaultextension = f".{fmt}",
            filetypes        = ftypes + [("All Files", "*.*")],
            initialfile      = default_name,
        )
        if not path_str:
            return

        try:
            p = Path(path_str)
            if fmt == "xlsx":
                n = self._store.export_xlsx(p)
            elif fmt == "txt":
                n = self._store.export_txt(p)
            elif fmt == "json":
                n = self._store.export_json(p)
            else:
                raise ValueError(f"Unsupported export format: {fmt!r}")
            _toast(
                self,
                f"✔  Exported {n} item{'s' if n != 1 else ''} → {p.name}",
                kind="success",
                duration=4_500,
            )
        except Exception as exc:
            log.error("Export failed (%s): %s\n%s", fmt.upper(), exc, traceback.format_exc())
            _toast(self, f"Export failed: {exc}", kind="danger", duration=5_000)

    # ── Settings ──────────────────────────────────────────────────────────

    def _open_settings(self) -> None:
        SettingsDialog(self, self._settings, on_save=self._apply_settings)

    def _apply_settings(self) -> None:
        self._settings.save()
        self._store.update_settings(self._settings)
        self._rebuild_fonts()
        _toast(self, "Settings saved successfully", kind="success")

    # ── Shutdown ──────────────────────────────────────────────────────────

    def _on_close(self) -> None:
        """
        Ordered, non-blocking shutdown.

        FIX-C: Instead of calling self.after() from the shutdown daemon thread
        (unsafe), the daemon thread sets _store_done_event; the main thread
        polls it via _poll_for_destroy() scheduled with after().
        """
        log.info("Application shutdown initiated")

        # Step 1 — disable polling callbacks
        self._running = False

        # Step 1b — hide window immediately
        try:
            self.withdraw()
        except Exception:
            pass

        # Step 2 — cancel named recurring IDs
        for aid in (self._queue_poll_id, self._watchdog_id):
            if aid is not None:
                try:
                    self.after_cancel(aid)
                except Exception:
                    pass
        self._queue_poll_id = None
        self._watchdog_id   = None

        # Step 3 — cancel remaining one-shot IDs
        for aid in list(self._after_ids):
            try:
                self.after_cancel(aid)
            except Exception:
                pass
        self._after_ids.clear()

        # Step 4 — cancel search debounce
        if self._search_timer is not None:
            try:
                self.after_cancel(self._search_timer)
            except Exception:
                pass
            self._search_timer = None

        # Step 5 — stop hotkey listener (daemon=True ensures process exits
        # even if join times out)
        if self._hotkey_listener is not None:
            try:
                self._hotkey_listener.stop()
            except Exception:
                pass
            try:
                if hasattr(self._hotkey_listener, "join"):
                    self._hotkey_listener.join(timeout=2.0)
            except Exception:
                pass

        # Step 6 — stop clipboard monitor
        self._monitor.stop()

        # Step 6b — drain any items queued between _running=False and stop()
        drained = 0
        while True:
            try:
                kind, payload = self._ui_queue.get_nowait()
                drained += 1
                if kind == "add":
                    try:
                        self._store.add(payload)
                    except Exception:
                        pass
            except queue.Empty:
                break
        if drained:
            log.info("Shutdown: drained %d pending clipboard item(s)", drained)

        # Step 7 — persist store on background thread; signal main thread when done
        def _finish_shutdown() -> None:
            try:
                self._store.stop()
            except Exception as exc:
                log.error("Error during store shutdown: %s", exc)
            finally:
                # FIX-C: signal main thread via Event, not self.after()
                self._store_done_event.set()

        threading.Thread(
            target=_finish_shutdown, daemon=True, name="ClipMgr/Shutdown"
        ).start()

        # Step 8 — FIX-C: poll from main thread until store is done, then destroy
        self._poll_for_destroy()

    def _poll_for_destroy(self) -> None:
        """
        FIX-C: Main-thread polling loop. Checks whether the store shutdown
        thread has finished (via threading.Event). When done, calls destroy()
        safely on the Tk event loop, avoiding the unsafe cross-thread after().
        """
        if self._store_done_event.is_set():
            self._do_destroy()
        else:
            try:
                self.after(_SHUTDOWN_POLL_MS, self._poll_for_destroy)
            except Exception:
                # If after() fails the mainloop is probably gone; try destroy directly
                self._do_destroy()

    def _do_destroy(self) -> None:
        """Final Tk teardown, called on the main thread."""
        try:
            self.destroy()
        except Exception:
            pass
        log.info("Application shutdown complete")

    # ── Global exception hook ─────────────────────────────────────────────

    def report_callback_exception(self, exc_type, exc_val, exc_tb) -> None:
        """Catch and log all unhandled Tk callback exceptions gracefully."""
        msg = "".join(traceback.format_exception(exc_type, exc_val, exc_tb))
        log.critical("Unhandled UI exception:\n%s", msg)
        try:
            _toast(
                self,
                "An unexpected error occurred — see log file for details.",
                kind="danger",
                duration=5_000,
            )
        except Exception:
            pass


# ═════════════════════════════════════════════════════════════════════════════
# ENTRY POINT
# ═════════════════════════════════════════════════════════════════════════════

def main() -> None:
    """
    Application entry point.
    Sets a global exception hook, starts the Tk main loop, and handles
    KeyboardInterrupt and fatal exceptions cleanly.
    """
    def _global_excepthook(exc_type, exc_val, exc_tb):
        log.critical(
            "Unhandled exception at top level:\n%s",
            "".join(traceback.format_exception(exc_type, exc_val, exc_tb)),
        )

    sys.excepthook = _global_excepthook

    try:
        app = ClipboardManagerApp()
        app.mainloop()
    except KeyboardInterrupt:
        log.info("Application interrupted by user (Ctrl+C)")
    except Exception:
        log.critical("Fatal error during startup or main loop:\n%s", traceback.format_exc())
        sys.exit(1)


if __name__ == "__main__":
    main()
